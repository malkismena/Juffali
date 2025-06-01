from langchain_google_genai import ChatGoogleGenerativeAI
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

import os
from dotenv import load_dotenv

load_dotenv()
API_KEY = os.getenv("API_KEY")
# Initialize Gemini
llm = ChatGoogleGenerativeAI(
    google_api_key=API_KEY,
    model="gemini-2.0-flash-exp",
    temperature=0.1,
    max_output_tokens=300
)

# Known Excel files mapped from txt names
known_files = {
    "تقرير شهر مايو.txt": "تقرير شهر مايو.xlsx",
    "تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.txt": "تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.xlsx"
}

# Gemini prompt
PROMPT = """
You are a document classifier and information extractor.

You will receive an Arabic input that may be structured or unstructured. You must:
1. Classify whether it belongs to:
   - "تقرير شهر مايو.txt" → Training Report (fields: الاسم، الهوية، الميلاد، المجال التدريبي، الهدف التدريبي، التقييم، الملاحظات)
   - "تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.txt" → Activities (fields: الاسم، رياضي، ترفيهي، اجتماعي، سباحة، الفئات، الملاحظات)

2. Extract the relevant structured fields regardless of format or order.

📌 Example 1 (Training Report):
Input:
عمر ماجد من مواليد 2006-02-01، رقمه 1234567890، يتدرب على الاستقلالي وهدفه ارتداء الملابس، حصل على 85 في التقييم وكان تحسنه ملحوظ.

Output:
{
  "target_datasheet": "تقرير شهر مايو.txt",
  "reason": "النص يحتوي على تقييم وهدف تدريبي مما يشير لتقرير تدريبي.",
  "structured_data": {
    "name": "عمر ماجد",
    "id": "1234567890",
    "dob": "2006-02-01",
    "training_field": "الاستقلالي",
    "training_goal": "ارتداء الملابس",
    "goal_evaluation": 85,
    "notes": "تحسنه ملحوظ"
  }
}

📌 Example 2 (Activities):
Input:
الطالب سعود ماجد  يشارك في تمارين التوازن، ويفضل الألعاب التي تستخدم إشارات مرئية. يظهر استجابة أفضل عند تقديم الأوامر عبر الوسائل البصريةو  يتفاعل مع مجموعة صغيرة. يفضل التفاعل مع عدد محدود من الأفراد
 . تصنيفه رياضي (ج) وترفيهي (ب) واجتماعي (ج) وسباحة (د).
Output:

{
  "target_datasheet": "تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.txt",
  "reason": "النص يحتوي على أنشطة وتصنيفات ترفيهية وسباحة مما يدل على ملف التصنيف.",
  "structured_data": {
  "name": "سعود ماجد ابراهيم العدوان",
  "activities": {
    "رياضي": {
      "category": "ج",
      "note": "يشارك في تمارين التوازن. يواجه صعوبة في متابعة التعليمات ويحتاج إلى إشراف مستمر."
    },
    "ترفيهي": {
      "category": "ب",
      "note": "يفضل الألعاب التي تستخدم إشارات مرئية. يظهر استجابة أفضل عند تقديم الأوامر عبر الوسائل البصرية."
    },
    "اجتماعي": {
      "category": "ج",
      "note": "يتفاعل مع مجموعة صغيرة. يفضل التفاعل مع عدد محدود من الأفراد."
    },
    "سباحة": {
      "category": "د",
      "note": "يحتاج إلى دعم مادي في المياه. يحتاج إشراف دائم في بيئة السباحة."
    }
  }
}

Now process this input and return only valid JSON:

\"\"\"{input_text}\"\"\"
"""

# 🔍 Function to classify and extract data from input text
def classify_input(query: str) -> str:
    try:
        full_prompt = f"{PROMPT}{query}"
        response = llm.invoke(full_prompt)

        if not response or not response.content:
            return "No classification available. Please try again."

        return response.content.strip()

    except Exception as e:
        return f"An error occurred during classification: {e}"

# 💾 Function to save extracted data to the actual Excel file
def save_structured_data_to_existing_sheet(data_json: str):
    try:
        # 🔧 Clean Gemini markdown formatting
        cleaned_json = data_json.strip()
        if cleaned_json.startswith("```json"):
            cleaned_json = cleaned_json[7:].strip()
        if cleaned_json.endswith("```"):
            cleaned_json = cleaned_json[:-3].strip()

        open_braces = cleaned_json.count('{')
        close_braces = cleaned_json.count('}')
        if open_braces > close_braces:
            cleaned_json += '}' * (open_braces - close_braces)
            
        data = json.loads(cleaned_json)
        target_file_txt = data.get("target_datasheet")
        structured_data = data.get("structured_data")

        if not target_file_txt or not structured_data:
            return "Invalid Gemini output: missing target_datasheet or structured_data."

        known_files = {
            "تقرير شهر مايو.txt": "تقرير شهر مايو.xlsx",
            "تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.txt": "تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.xlsx"
        }

        if target_file_txt=="تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.txt":
            target_file_xlsx = known_files.get(target_file_txt)
            wb = load_workbook(target_file_xlsx) 

            ws = wb.active  
              
            def find_next_empty_row(ws, col_letter="F", start_row=11):
                row = start_row
                while ws[f"{col_letter}{row}"].value:
                    row += 1
                return row

            row = find_next_empty_row(ws)

            activity_keys = ["رياضي", "ترفيهي", "اجتماعي", "سباحة"]
            columns = ["G", "H", "I", "J"]

            for col, key in zip(columns, activity_keys):
                category = structured_data.get("activities", {}).get(key, {}).get("category", "")
                ws[f"{col}{row}"] = category

            ws[f"F{row}"] = structured_data.get("name", "")

            
            notes_list = []
            for key in activity_keys:
                note = structured_data.get("activities", {}).get(key, {}).get("note")
                if note:
                    notes_list.append(f"- {key}: {note}")

            combined_notes = "\n".join(notes_list)

            cell = ws[f"K{row}"]
            cell.value = combined_notes

            #cell.font = Font(size=36) 
            ws.column_dimensions["K"].width = 200  
            ws.row_dimensions[row].height = 300  


        elif target_file_txt=="تقرير شهر مايو.txt":
            target_file_xlsx = known_files.get(target_file_txt)
            wb = load_workbook(target_file_xlsx)
            ws = wb.active

            def find_next_empty_row(ws, col_letter="C", start_row=7):
                row = start_row
                while ws[f"{col_letter}{row}"].value:
                    row += 4 
                return row

            row = find_next_empty_row(ws)
            ws[f"C{row}"] = structured_data.get("name", "")
            ws[f"D{row}"] = structured_data.get("id", "")
            ws[f"E{row}"] = structured_data.get("dob", "")
            field_row_map = {
                "الاستقلالي": 0,
                "الإدراكي": 1,
                "الاجتماعي": 2,
                "التواصل": 3
            }

            field = structured_data.get("training_field")
            if field in field_row_map:
                offset = field_row_map[field]
                target_row = row + offset

                ws[f"G{target_row}"] = structured_data.get("training_goal", "")
                

                evaluation = structured_data.get("goal_evaluation")

                if evaluation not in (None, ""):
                    try:
                        evaluation = int(evaluation)  

                        if evaluation <= 60:
                            ws[f"J{target_row}"] = evaluation
                        elif 80 <= evaluation < 100:
                            ws[f"I{target_row}"] = evaluation
                        elif evaluation >= 100:
                            ws[f"H{target_row}"] = evaluation

                    except (ValueError, TypeError):
                        pass

                ws[f"K{target_row}"] = structured_data.get("notes", "")

        else:
            return f"❌ Sheet for {target_file_txt} not found in current directory."
        

        wb.save(target_file_xlsx)
        print("✅")
        return f"✅ Data saved to existing file: {target_file_xlsx}"
    

    except Exception as e:
        return f"❌ Error: {e}"

# 🚀 Main Execution
def main(user_input):
    #user_input = input("Enter any sentence in Arabic (structured or unstructured):\n")
    class_file = classify_input(user_input)

    print("\n🔤 Gemini Response:")
    print(class_file)

    print("\n💾 Saving to Excel...")
    result = save_structured_data_to_existing_sheet(class_file)
    print(result)
    return result
