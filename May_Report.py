
from openpyxl import load_workbook

# ملف الاكسل الموجود
excel_file = "تقرير شهر مايو.xlsx"

# افتح ملف الاكسل (load)
wb = load_workbook(excel_file)

# اختر الورقة النشطة أو بالاسم
ws = wb.active  # أو مثلاً wb["Sheet1"] إذا تعرف اسم الورقة

# بياناتك المستخرجة من JSON
structured_data = {
    "name": "عبدالله بن راشد عبدالله بو حسن",
    "id": "1056783005",
    "dob": "2006-10-11",
    "training_field": "الإدراكي",
    "training_goal": "استخدام دورة المياه بشكل مستقل",
    "goal_evaluation": 90,
    "notes": "يحتاج خطة دعم منزلي"
}
def find_next_empty_row(ws, col_letter="C", start_row=7):
    row = start_row
    while ws[f"{col_letter}{row}"].value:
        row += 4 
    return row

row = find_next_empty_row(ws)
ws[f"C{row}"] = structured_data.get("name")
ws[f"D{row}"] = structured_data.get("id")
ws[f"E{row}"] = structured_data.get("dob")
field_row_map = {
    "الاستقلالي": 0,
    "الإدراكي": 1,
    "الاجتماعي": 2,
    "التواصل": 3
}

field = structured_data.get("training_field")
if field in field_row_map:
    offset = field_row_map[field]
    target_row = row + offset

    ws[f"G{target_row}"] = structured_data.get("training_goal")
    
    evaluation = structured_data.get("goal_evaluation")
    if evaluation <= 60:
        ws[f"J{target_row}"] = evaluation
    elif evaluation >= 80:
        ws[f"H{target_row}"] = evaluation
    else:
        ws[f"I{target_row}"] = evaluation

    ws[f"K{target_row}"] = structured_data.get("notes")

# حفظ التعديلات في الملف
wb.save(excel_file)

print("✅ تم تحديث البيانات في الخلايا المحددة.")
