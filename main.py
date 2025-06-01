
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

excel_file = "تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.xlsx"

wb = load_workbook(excel_file)

ws = wb.active  
  
def find_next_empty_row(ws, col_letter="F", start_row=11):
    row = start_row
    while ws[f"{col_letter}{row}"].value:
        row += 1
    return row

row = find_next_empty_row(ws)
ws[f"F{row}"] = structured_data.get("name")
ws[f"G{row}"] = structured_data.get("activities", {}).get("رياضي", {}).get("category")
ws[f"H{row}"] = structured_data.get("activities", {}).get("ترفيهي", {}).get("category")
ws[f"I{row}"] = structured_data.get("activities", {}).get("اجتماعي", {}).get("category")
ws[f"J{row}"] =  structured_data.get("activities", {}).get("سباحة", {}).get("category")

notes_list = []
for activity in structured_data["activities"].values():
    note = activity.get("note")
    if note:
        notes_list.append(f"- {note}")

combined_notes = "\n".join(notes_list)

cell = ws[f"K{row}"]
cell.value = combined_notes
#cell.font = Font(size=36) 
ws.column_dimensions["K"].width = 200  
ws.row_dimensions[row].height = 300  

wb.save(excel_file)
print("✅")