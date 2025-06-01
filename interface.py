"""import streamlit as st
import os
from openpyxl import load_workbook
from translator import main  

st.set_page_config(page_title="Munazzim Chatbot", layout="wide")

# دالة التحقق من وجود بيانات (ممكن تستخدمها لاحقًا لو حبيت)
def has_data(filepath, data_column="C", start_row=7):
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        for row in range(start_row, ws.max_row + 1):
            if ws[f"{data_column}{row}"].value:
                return True
        return False
    except:
        return False

available_files = [
    ("تقرير شهر مايو.xlsx", "C", 7),
    ("تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.xlsx", "F", 11)
]

with st.sidebar:
    st.markdown("### 📝 مساعد إدخال البيانات")
    st.markdown("الرجاء إدخال المعلومات المطلوبة بدقة.\nسيتم حفظها تلقائيًا في ملف Excel.")

    st.markdown("---")
    st.markdown("### 📂 تحميل الملفات:")

    found = False
    for file_name, col, start_row in available_files:
        if os.path.exists(file_name):  # فقط تحقق من وجود الملف
            with open(file_name, "rb") as f:
                st.download_button(
                    label=f"📥 تحميل {file_name}",
                    data=f,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            found = True

    if not found:
        st.info("لا توجد ملفات للتحميل حالياً.")

st.markdown("<h1 style='text-align: center; color: #2c3e50;'>Munazzim Chatbot</h1>", unsafe_allow_html=True)

if "conversation_history" not in st.session_state:
    st.session_state.conversation_history = []

for message in st.session_state.conversation_history:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

user_input = st.chat_input("💬 أدخل الجملة (بالعربية، منظمة أو غير منظمة):")

if user_input:
    with st.chat_message("user"):
        st.markdown(user_input)

    st.session_state.conversation_history.append({
        "role": "user",
        "content": user_input,
    })

    ai_response = main(user_input)

    with st.chat_message("assistant"):
        st.markdown(ai_response)

    st.session_state.conversation_history.append({
        "role": "assistant",
        "content": ai_response,
    })
    """
import streamlit as st
import os
from translator import main
from io import BytesIO

# ملفات Excel المطلوبة
available_files = [
    ("تقرير شهر مايو.xlsx", "C", 7),
    ("تصنيف مستفيدي السكن الداخلي في الأنشطة الترفيهية والاجتماعية.xlsx", "F", 11)
]

# تحميل محتوى ملف كـ BytesIO
def load_file_bytes(file_path):
    with open(file_path, "rb") as f:
        return f.read()

st.set_page_config(page_title="Munazzim Chatbot", layout="wide")

st.markdown("<h1 style='text-align: center; color: #2c3e50;'>Munazzim Chatbot</h1>", unsafe_allow_html=True)

if "conversation_history" not in st.session_state:
    st.session_state.conversation_history = []

# 🟦 الشريط الجانبي
with st.sidebar:
    st.markdown("### 📝 مساعد إدخال البيانات")
    st.markdown("الرجاء إدخال المعلومات المطلوبة بدقة.\nسيتم حفظها تلقائيًا في ملف Excel.")
    st.markdown("---")

    st.markdown("### 📂 تحميل الملفات المحدثة:")

    for file_name, _, _ in available_files:
        if os.path.exists(file_name):
            file_bytes = load_file_bytes(file_name)
            st.download_button(
                label=f"📥 تحميل {file_name}",
                data=file_bytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"sidebar-{file_name}"
            )
        else:
            st.warning(f"⚠️ الملف غير موجود: {file_name}")

# 🗨️ عرض المحادثات السابقة
for message in st.session_state.conversation_history:
    with st.chat_message(message["role"]):
        st.markdown(message["content"])

# 📝 إدخال المستخدم
user_input = st.chat_input("💬 أدخل الجملة (بالعربية، منظمة أو غير منظمة):")

if user_input:
    # عرض رسالة المستخدم
    with st.chat_message("user"):
        st.markdown(user_input)

    st.session_state.conversation_history.append({
        "role": "user",
        "content": user_input,
    })

    # الرد من الذكاء الاصطناعي + حفظ البيانات
    ai_response = main(user_input)

    # عرض رد الذكاء الاصطناعي
    with st.chat_message("assistant"):
        st.markdown(ai_response)

    st.session_state.conversation_history.append({
        "role": "assistant",
        "content": ai_response,
    })

    # ✅ إعادة تحميل الأزرار المحدثة في الشريط الجانبي
    with st.sidebar:
        st.markdown("---")
        st.markdown("### 🔁 تحديث الملفات بعد الإدخال:")

        for file_name, _, _ in available_files:
            if os.path.exists(file_name):
                file_bytes = load_file_bytes(file_name)
                st.download_button(
                    label=f"📥 تحميل {file_name} (مُحدث)",
                    data=file_bytes,
                    file_name=file_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"sidebar-update-{file_name}"
                )
